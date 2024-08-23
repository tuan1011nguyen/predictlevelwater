#import openpyxl
import numpy as np
from sklearn.metrics import r2_score
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_squared_error
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def NSE(y_test, y_pred):
    return (1 - (np.sum((y_pred - y_test) ** 2) / np.sum((y_test - np.mean(y_test)) ** 2)))

def R2(y_test, y_pred):
    y_test_mean = np.mean(y_test)
    #print(y_test_mean)
    y_pred_mean = np.mean(y_pred)
    #print(y_pred_mean)
    S_res = np.sum((y_test - y_test_mean)*(y_pred - y_pred_mean))
    #print(S_res)
    S_tot1 = np.sum((y_test - y_test_mean)**2)
    #print(S_tot1)
    S_tot2 = np.sum((y_pred - y_pred_mean)**2)
    #print(S_tot2)
    S_tot = np.sqrt(S_tot1 * S_tot2)
    #print(S_tot1 * S_tot2, S_tot)
    return (S_res / S_tot) ** 2

def MAE(y_test, y_pred):
    return mean_absolute_error(y_test, y_pred)

def RMSE(y_test, y_pred):
    return mean_squared_error(y_test, y_pred, squared=False)

def MAX_ERROR(y_test, y_pred):
    max = 0
    for i in range(0, len(y_test)):
        if (max < abs(y_test[i] - y_pred[i])):
            max = abs(y_test[i] - y_pred[i])
    return max

def OTR(y_test, y_pred, threshold):
    r = 0
    for i in range(0, len(y_test)):
        if (threshold < abs(y_test[i] - y_pred[i])):
            r += 1
    return r/len(y_test)

def output_Excel(X_test, y_test, y_pred, attributes,know_attributes, pred_attribute, numdays, afterdays, name_fig, output_excel_path, threshold, name_ref_column, test_ref_column):
    # Xác định số hàng và cột lớn nhất trong file excel cần tạo
    num_rows = len(X_test)
    columnX_test = len(X_test[0])
    num_attributes = len(attributes)
    num_know_attributes = len(know_attributes)

    # Tạo một workbook mới và active nó
    wb = Workbook()
    ws = wb.active

    # Ghi cột tham chiếu vào file excel
    if (len(name_ref_column)==0):
        ws.cell(column=1, row=1, value='STT')
        for i in range(num_rows):
            ws.cell(column=1, row=i + 2, value=i+1)
    else:
        ws.cell(column=1, row=1, value=name_ref_column[0])
        for i in range(len(test_ref_column)):
            ws.cell(column=1, row=i+2, value=test_ref_column[i])
    # Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
    for j in range(numdays):
        for k in range(num_attributes):
            v = attributes[k] + str('_') + str(j + 1)
            ws.cell(column=num_attributes*j + k + 2, row=1, value=v)
    if(num_know_attributes > 0):
        for j in range(afterdays):
            for k in range(num_know_attributes):
                v = know_attributes[k] + str('_') + str(numdays+j+1)
                ws.cell(column=numdays * num_attributes + num_know_attributes * j + k+2, row=1, value=v)
    v = pred_attribute[0] + str(afterdays) + '_real'
    ws.cell(column=columnX_test+2, row=1, value=v)
    v = pred_attribute[0] + str(afterdays) + '_prediction'
    ws.cell(column=columnX_test + 3, row=1, value=v)
    ws.cell(column=columnX_test + 4, row=1, value='ASB(real_value, prediction_value')
    for i in range(0, num_rows):
        for j in range(0, columnX_test):
            v = X_test[i][j]
            ws.cell(column=j+2, row=i+2, value=v)

    for i in range(0, num_rows):
        v = y_test[i]
        ws.cell(column=columnX_test + 2, row=i + 2, value=v)
        v = y_pred[i]
        ws.cell(column=columnX_test + 3, row=i + 2, value=v)
        v = abs(y_test[i]-y_pred[i])
        ws.cell(column=columnX_test + 4, row=i + 2, value=v)
    #In độ đo NSE
    ws.cell(column=1, row=num_rows + 2, value='Nash-Sutcliffe efficiency (NSE)')
    ws.cell(column=2, row=num_rows + 2, value=NSE(y_test, y_pred))
    #In độ đo R2
    ws.cell(column=1, row=num_rows + 3, value='Coefficient of determination (R2)')
    ws.cell(column=2, row=num_rows + 3, value=R2(y_test, y_pred))
    # In độ đo MAE
    ws.cell(column=1, row=num_rows + 4, value='Mean absolute error (MAE)')
    ws.cell(column=2, row=num_rows + 4, value=MAE(y_test, y_pred))
    # In độ đo RMSE
    ws.cell(column=1, row=num_rows + 5, value='Root mean square error (RMSE)')
    ws.cell(column=2, row=num_rows + 5, value=RMSE(y_test, y_pred))
    #in độ đo OTR
    ws.cell(column=1, row=num_rows + 6, value='OTR')
    ws.cell(column=2, row=num_rows + 6, value=OTR(y_test, y_pred, threshold))

    # In độ đo MAX error
    ws.cell(column=1, row=num_rows + 7, value='MAX error')
    ws.cell(column=2, row=num_rows + 7, value=MAX_ERROR(y_test, y_pred))

    #Chèn ảnh
    img = Image(name_fig)
    img.width, img.height = 1200, 600
    cell_img = 'C' + str(num_rows + 10)
    img.anchor = cell_img
    ws.add_image(img)

    # Lưu lại file Excel
    wb.save(output_excel_path)