import numpy as np
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.linear_model import LinearRegression, Ridge
from sklearn.neural_network import MLPRegressor
from read_data import get_data
from write_data import output_Excel
from matplotlib import pyplot
from openpyxl import Workbook
import write_data
import time
import os
from tensorflow.keras import Sequential
from tensorflow.keras.layers import Dense, BatchNormalization, Dropout
from tensorflow.keras.layers import LSTM
from sklearn.model_selection import train_test_split

from attention import attention
from keras.layers import LSTM, TimeDistributed, RepeatVector, Layer

def LSTM_Model(file_train, file_test, know_attributes, unknow_attributes, pred_attribute, foldername, max_numdays, max_afterdays, threshold, nomalize, smote, smote_threshold, name_ref_column):
    if not os.path.exists(foldername):
        os.makedirs(foldername)

    #Các file kết quả sẽ xuất theo tên của thư mục con này.
    foldername_split = foldername.split("/")
    folderchild = foldername_split[-1]
    Sumary_result_file = foldername + '/Sumary_result_' + folderchild + '.xlsx'
    feature_attributes = know_attributes + unknow_attributes
    attributes=[]
    for i in range(len(feature_attributes)):
        if feature_attributes[i] not in pred_attribute:
            attributes.append(feature_attributes[i])
    for i in range(len(feature_attributes)):
        if feature_attributes[i] in pred_attribute:
            attributes.append(feature_attributes[i])
    print('attributes:', attributes)
    all_attributes = []
    for i in range(len(attributes)):
        if attributes[i] not in pred_attribute:
            all_attributes.append(attributes[i])
    for i in range(len(pred_attribute)):
        all_attributes.append(pred_attribute[i])
    print('all_attributes:', all_attributes)


    wb = Workbook()
    ws = wb.active
    ws.cell(column=1, row=1, value='Predict Time')
    ws.cell(column=2, row=1, value='Back Time')
    ws.cell(column=3, row=1, value='NSE')
    ws.cell(column=4, row=1, value='R2')
    ws.cell(column=5, row=1, value='MAE')
    ws.cell(column=6, row=1, value='RMSE')
    ws.cell(column=7, row=1, value='OTR')
    ws.cell(column=8, row=1, value='MAX error')
    ws.cell(column=9, row=1, value='Time (sec)')

    ''''if (len(know_attributes)==0):
        start_index = 1
        max_numdays += 1
    else:
        start_index = 1'''''

    normalize = False
    #numdays là số ngày dùng để dự đoán
    #afterdays là số ngày muốn dự đoán
    count = 1
    for afterdays in range(max_afterdays, max_afterdays + 1):
        for numdays in range(1, max_numdays + 1):
            count += 1
            #numdays1 = numdays-(start_index-1)
            filename = foldername + '/result_' + folderchild + '_predictime_' + str(afterdays) + '_backtime_'+ str(numdays)
            output_file = filename + '.xlsx'
            start_time = time.time()
            X_test_original, y_test_original, X_train, y_train, X_test, y_test, scalerX, scalerY, test_ref_column = get_data(file_train, file_test,
                                                                                                        all_attributes,
                                                                                                        attributes,
                                                                                                        know_attributes,
                                                                                                        numdays, afterdays,
                                                                                                     normalize, smote, smote_threshold, name_ref_column)
            if (len(know_attributes) == 0):
                num_attributes = int(X_train.shape[1]/numdays)
                X_LSTM = X_train.reshape(X_train.shape[0], numdays, num_attributes)
                X_test_LSTM = X_test.reshape(X_test.shape[0], numdays, num_attributes)
            else:
                X_LSTM = X_train.reshape(X_train.shape[0], 1, X_train.shape[1])
                X_test_LSTM = X_test.reshape(X_test.shape[0], 1, X_test.shape[1])

            print('X_LSTM', X_LSTM.shape)

            X_train_LSTM, X_val_LSTM, y_train_LSTM, y_val_LSTM = train_test_split(X_LSTM, y_train, test_size=0.2, shuffle = True)

            model = Sequential()
            model.add(LSTM(128, activation='tanh', input_shape=(X_train_LSTM.shape[1], X_train_LSTM.shape[2]), return_sequences = True))
            model.add(attention())
            model.add(RepeatVector(X_train_LSTM.shape[1]))
            model.add(LSTM(128, activation='tanh', input_shape=(X_train_LSTM.shape[1], X_train_LSTM.shape[2]), return_sequences = False))
    
            #model.add(BatchNormalization())
            #model.add(Dropout(0.2))
            #model.add(Dense(1000, activation='tanh'))
            #model.add(BatchNormalization())
            #model.add(Dropout(0.2))
            #model.add(Dense(800, activation='tanh'))
            #model.add(BatchNormalization())
            #model.add(Dense(1000, activation='tanh'))
            #model.add(BatchNormalization())
            #model.add(Dense(150, activation='tanh'))
            model.add(Dense(1))
            # compile the model
            model.compile(optimizer='adam', loss='mse')
            # fit the model
            model.fit(X_train_LSTM, y_train_LSTM, epochs=2, batch_size=30, verbose=1, validation_data=(X_val_LSTM, y_val_LSTM))
            # evaluate the model
            y_pred_scaler = model.predict(X_test_LSTM)
            y_pred_scaler = y_pred_scaler.flatten()

            if normalize == False:
                y_pred = y_pred_scaler
            else:
                y_test = y_test_original
                y_pred_scaler = y_pred_scaler.reshape(len(y_pred_scaler), 1)
                y_pred_2D = scalerY.inverse_transform(y_pred_scaler)
                #y_pred_2D = np.array(y_pred_2D)
                y_pred = y_pred_2D.flatten()

            end_time = time.time()

            fig, ax = pyplot.subplots(figsize=(23, 5))
            ax.plot(y_pred, color='red', label='y_pred')
            ax.plot(y_test, color='black', label='y_test')
            ax.legend()
            name_fig = filename + '.png'
            fig.savefig(name_fig, format='png')
            pyplot.close(fig)
            output_Excel(X_test_original, y_test_original, y_pred, attributes, know_attributes, pred_attribute, numdays, afterdays, name_fig, output_file, threshold, name_ref_column, test_ref_column)

            ws.cell(column=1, row=count, value=afterdays)
            ws.cell(column=2, row=count, value=numdays)
            ws.cell(column=3, row=count, value=write_data.NSE(y_test, y_pred))
            ws.cell(column=4, row=count, value=write_data.R2(y_test, y_pred))
            ws.cell(column=5, row=count, value=write_data.MAE(y_test, y_pred))
            ws.cell(column=6, row=count, value=write_data.RMSE(y_test, y_pred))
            ws.cell(column=7, row=count, value=write_data.OTR(y_test, y_pred, threshold))
            ws.cell(column=8, row=count, value=write_data.MAX_ERROR(y_test, y_pred))
            ws.cell(column=9, row=count, value=end_time - start_time)

    wb.save(Sumary_result_file)
    print('Well done!')
#LR_Model(file_train, file_test, know_attributes, unknow_attributes, pred_attribute, foldername, max_numdays, max_afterdays, nomalize=False)